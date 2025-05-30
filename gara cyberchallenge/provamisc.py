import numpy as np
from openpyxl import load_workbook
import re

def parse_cell_ref(s):
    col_part = ''.join(filter(str.isalpha, s))
    row_part = ''.join(filter(str.isdigit, s))
    if not col_part or not row_part:
        raise ValueError(f"Invalid cell reference: {s}")
    row_idx = int(row_part) - 1
    col_idx = 0
    for char in col_part:
        col_idx = col_idx * 26 + (ord(char.upper()) - ord('A') + 1)
    col_idx -= 1
    return (row_idx, col_idx)

def extract_conditions_from_and_formula(formula):
    formula = formula.strip()
    if not formula.upper().startswith('=AND('):
        raise ValueError("Invalid AND formula: must start with '=AND('")
    if not formula.endswith(')'):
        raise ValueError("Invalid AND formula: must end with ')'")
    inner = formula[formula.find('(')+1:-1]
    parts = [p.strip() for p in inner.split(',')]
    cell_refs = []
    for part in parts:
        if '=' in part:
            cell_ref = part.split('=')[0].strip()
            cell_refs.append(cell_ref)
        else:
            cell_refs.append(part)
    return cell_refs

def parse_formula(s, get_expr_func):
    if not s.startswith('='):
        s = '=' + s
    s = s[1:]
    if re.match(r'^-?\d+$', s):
        return (int(s), [0] * 30)
    m = re.match(r'^(-?\d+)\*([A-Z]+\d+)$', s)
    if m:
        constant = int(m.group(1))
        target_cell = m.group(2)
        t_row, t_col = parse_cell_ref(target_cell)
        c0, coeffs0 = get_expr_func(t_row, t_col)
        new_constant = constant * c0
        new_coeffs = [constant * c for c in coeffs0]
        return (new_constant, new_coeffs)
    m = re.match(r'^([+\-]?[A-Z]+\d+)([+\-])([A-Z]+\d+)$', s)
    if m:
        token1 = m.group(1)
        op = m.group(2)
        token2 = m.group(3)
        sign1 = 1
        if token1.startswith('-'):
            sign1 = -1
            token1 = token1[1:]
        elif token1.startswith('+'):
            token1 = token1[1:]
        t1_row, t1_col = parse_cell_ref(token1)
        c1, coeffs1 = get_expr_func(t1_row, t1_col)
        t2_row, t2_col = parse_cell_ref(token2)
        c2, coeffs2 = get_expr_func(t2_row, t2_col)
        factor2 = 1 if op == '+' else -1
        new_constant = sign1 * c1 + factor2 * c2
        new_coeffs = [sign1 * a + factor2 * b for a, b in zip(coeffs1, coeffs2)]
        return (new_constant, new_coeffs)
    m = re.match(r'^([A-Z]+\d+)$', s)
    if m:
        target_cell = m.group(1)
        t_row, t_col = parse_cell_ref(target_cell)
        return get_expr_func(t_row, t_col)
    raise ValueError(f"Unsupported formula: {s}")

def main():
    wb = load_workbook("/home/mattia/Downloads/Book1.xlsx", data_only=False)
    ws = wb.active  # o specifica il foglio, es. wb["Foglio1"]

    max_rows = ws.max_row
    max_cols = ws.max_column

    # Legge tutte le formule (o valori se non formula) e li mette in formulas[row][col]
    formulas = [['' for _ in range(max_cols)] for _ in range(max_rows)]

    for r in range(1, max_rows+1):
        for c in range(1, max_cols+1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None:
                val = ''
            if isinstance(val, str) and val.startswith('='):
                formulas[r-1][c-1] = val
            else:
                formulas[r-1][c-1] = str(val)

    # Stampa controllo prime righe e colonne
    print("Prime 5 righe e 5 colonne di formule:")
    for i in range(min(5, max_rows)):
        print(f"Row {i}: {formulas[i][:5]}")

    a2_formula = formulas[1][0]  # A2 è (1,0) zero based
    print("Formula in A2:", repr(a2_formula))

    try:
        condition_cell_refs = extract_conditions_from_and_formula(a2_formula)
    except Exception as e:
        print(f"Errore nell'estrazione delle condizioni dalla formula A2: {repr(a2_formula)}")
        raise e

    cache = {}

    def get_expr(row_idx, col_idx):
        key = (row_idx, col_idx)
        if key in cache:
            return cache[key]

        if row_idx == 2:
            coeffs = [0] * max_cols
            coeffs[col_idx] = 1
            cache[key] = (0, coeffs)
            return (0, coeffs)

        formula_str = ''
        if row_idx < len(formulas) and col_idx < len(formulas[row_idx]):
            formula_str = formulas[row_idx][col_idx]

        if not formula_str or formula_str.strip() == '' or formula_str.strip().lower() == 'x':
            cache[key] = (0, [0] * max_cols)
            return (0, [0] * max_cols)

        try:
            expr = parse_formula(formula_str, get_expr)
            cache[key] = expr
            return expr
        except Exception as e:
            print(f"Errore nella cella ({row_idx+1},{col_idx+1}) con formula: {repr(formula_str)}")
            raise e

    A = []
    b = []

    for cell_ref in condition_cell_refs:
        row_idx, col_idx = parse_cell_ref(cell_ref)
        const, coeffs = get_expr(row_idx, col_idx)
        A.append(coeffs)
        b.append(-const)

    import numpy as np
    A = np.array(A, dtype=np.float64)
    b = np.array(b, dtype=np.float64)

    x, residuals, rank, s = np.linalg.lstsq(A, b, rcond=None)
    x_int = []
    for num in x:
        # Mappa i valori in [0, 127] (ASCII stampabili) per sicurezza
        val = int(round(num)) % 128
        x_int.append(val)

    flag = ''.join(chr(c) for c in x_int)
    print(f"Flag: {flag}")

if __name__ == "__main__":
    main()
